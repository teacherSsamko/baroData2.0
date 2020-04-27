from datetime import datetime, timedelta, date

from openpyxl import Workbook, load_workbook

class BaroRoll:
    def __init__(self, wb):
        self.wb = wb
        self.sheet = wb[wb.sheetnames[0]]
        self.count_students = self.sheet.max_row - 1
        self.count_subjects = self.sheet.max_column - 1
        self.subjects = []
        self.make_subjects()
        self.data = [['이름', '과목', '이수 시간']]
        self.make_data()
        self.reordered = [['이름', '과목', '이수시간', '소요시간']]
        self.reorder()

    def make_subjects(self):
        for i in range(self.count_subjects):
            subject = self.sheet.cell(row=1, column=i + 2).value
            subject = subject.split("-")[1]
            self.subjects.append(subject)
        
    def make_data(self):
        for student in range(2, self.count_students + 2):
            for subject in range(2, self.count_subjects + 2):
                s_name = self.sheet.cell(row=student, column=1).value
                subj = self.subjects[subject - 2]
                # 이수 시간은 datetime 으로 저장
                checked_time = self.sheet.cell(
                    row=student, column=subject).value
                if checked_time == None:
                    checked_time = '결석'
                else:
                    checked_time = checked_time.rstrip()
                    checked_time = datetime.strptime(checked_time, "%m/%d %H:%M")
                    # zero leading datetime
                    checked_time = checked_time.strftime("%m/%d %H:%M")
                self.data.append([s_name, subj, checked_time])
    
    def reorder(self):
        temp = []
        for i in range(len(self.data)):
            if i == 0:
                continue
            if not i % self.count_subjects == 0:
                temp.append(self.data[i])
            else:
                temp.append(self.data[i])
                # 이수 시간 기준으로 정렬
                temp.sort(key=lambda temp: temp[2])
                # 소요시간 계산을 위한 temp2
                temp2 = [temp[0]] + temp[:-1]

                for i in range(len(temp)):
                    if temp[i][2] != '결석':
                        elapsed = datetime.strptime(
                            temp[i][2], "%m/%d %H:%M") - datetime.strptime(temp2[i][2], "%m/%d %H:%M")

                        if elapsed < timedelta(seconds=1):
                            temp[i].append('시작')
                            continue
                        elif elapsed < timedelta(hours=24):
                            dt = datetime(2020, 1, 1, 0, 0, 0) + elapsed
                            elapsed = f'({dt.strftime("%H:%M")})'
                        else:
                            elapsed = '(다른 날 수강)'
                    else:
                        elapsed = '-'
                    temp[i].append(elapsed)
                for item in temp:
                    self.reordered.append(item)
                temp = []
