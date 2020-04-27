import os
import locale
from os import listdir
from datetime import datetime, timedelta, date
from unicodedata import normalize

from openpyxl import Workbook, load_workbook
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.styles.differential import DifferentialStyle


from roll import BaroRoll

print('*'*50)
print("출결관리 프로그램v2.0 \n장애발생 시 삼양초 '이은섭'으로 메세지 주세요. :)")
print('*'*50)

BASE_DIR = os.path.dirname(os.path.realpath(__file__))
LIST_DIR = os.path.join(BASE_DIR, "roll_list/")
RESULT_DIR = os.path.join(BASE_DIR, "results")

list_files = listdir("roll_list")
list_files.sort()
print(list_files)

days = []
days_forsheet = []
days_i = 0

# 임시 결과 파일
temp_result = Workbook()
temp_result.save(os.path.join(RESULT_DIR, 'temp.xlsx'))

# 조건부 서식 입히기
# 소요 시간 3분 이내인 셀 > 표시
# red_fill = PatternFill(bgColor="FFC7CE")
# dxf = DifferentialStyle(fill=red_fill)
# r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
# r.formula = ['$D2>"(00:03)"']


for f in list_files:
    # 파일 읽기
    # 실행 중 생성되는 임시파일 무시
    if not '.xlsx' in f:
        continue
    # 파일 이름에서 날짜만 추출
    thisDay = f.split("-")[-1].rstrip('.xlsx')
    thisDay = normalize('NFC', thisDay)
    thisDay = datetime.strptime(thisDay, "%m월%d일")
    thisDay = thisDay.replace(year=2020)

    # 날짜 목록에 추가하기
    days.append(thisDay.strftime("%m%d"))
    days_forsheet.append(days[-1])

    thisDay = thisDay.strftime("%m월%d일(%a)")

    # 파일 이름에서 prefix 생성
    school_infos = f.split('-')[:-1]

    print('[처리중인 파일]')
    print(f)

    # 워크북 생성
    read_f = os.path.join(LIST_DIR, f)
    wb = load_workbook(read_f, read_only=True)

    # BaroRoll 인스턴스 생성
    baro = BaroRoll(wb)

    sheet = baro.sheet

    # 학생 수, 과목 수
    print('학생 수:', baro.count_students)
    print('과목 수:', baro.count_subjects)

    # 과목 리스트
    subjects = []
    for i in range(baro.count_subjects):
        subject = sheet.cell(row=1, column=i + 2).value
        subject = subject.split("-")[1]
        subjects.append(subject)


    print(subjects)

    # 데이터 가공
    data = [['이름', '과목', '이수 시간']]
    for student in range(2, baro.count_students + 2):
        for subject in range(2, baro.count_subjects + 2):
            s_name = sheet.cell(row=student, column=1).value
            subj = subjects[subject - 2]
            # 이수 시간은 datetime 으로 저장
            checked_time = sheet.cell(
                row=student, column=subject).value
            if checked_time == None:
                checked_time = '결석'
            else:
                checked_time = checked_time.rstrip()
                checked_time = datetime.strptime(checked_time, "%m/%d %H:%M")
                # zero leading datetime
                checked_time = checked_time.strftime("%m/%d %H:%M")
            data.append([s_name, subj, checked_time])

    # 이수 시각 기준으로 재정렬
    reordered = [['이름', '과목', '이수시간', '소요시간']]
    # 학생별로 항목을 모은 후 정렬하기 위한 temp
    temp = []
    for i in range(len(data)):
        if i == 0:
            continue
        if not i % baro.count_subjects == 0:
            temp.append(data[i])
        else:
            temp.append(data[i])
            # 이수 시간 기준으로 정렬
            temp.sort(key=lambda temp: temp[2])
            # 소요시간 계산을 위한 temp2
            temp2 = [temp[0]] + temp[:-1]

            for i in range(len(temp)):
                if temp[i][2] != '결석':
                    elapsed = datetime.strptime(
                        temp[i][2], "%m/%d %H:%M") - datetime.strptime(temp2[i][2], "%m/%d %H:%M")

                    if elapsed < timedelta(seconds=1):
                        temp[i].append('시작')
                        continue
                    elif elapsed < timedelta(hours=24):
                        dt = datetime(2020, 1, 1, 0, 0, 0) + elapsed
                        elapsed = f'({dt.strftime("%H:%M")})'
                    else:
                        elapsed = '(다른 날 수강)'
                else:
                    elapsed = '-'
                temp[i].append(elapsed)
            for item in temp:
                reordered.append(item)
            temp = []

    # 워크시트 작성

    result_wb = load_workbook(os.path.join(RESULT_DIR, "temp.xlsx"))
    if days_i == 0:
        days_forsheet[0] = result_wb.active
        days_forsheet[0].title = thisDay
    else:
        days_forsheet[days_i] = result_wb.create_sheet(
            title=thisDay)
    for row in reordered:
        days_forsheet[days_i].append(row)
    # total row
    # count_rows = days_forsheet[days_i].max_row
    # cell styling
    # days_forsheet[days_i].conditional_formatting.add(f"A1:D{count_rows}", r)
    days_i += 1
    result_wb.save(os.path.join(RESULT_DIR, "temp.xlsx"))


# 임시 결과 파일 이름 변경
tempfilename = os.path.join(RESULT_DIR, "temp.xlsx")
file_prefix = '_'.join(school_infos)
file_period = f'{days[0]}~{days[-1]}'

result_filename = os.path.join(
    RESULT_DIR, f'(출결){file_prefix}_{file_period}.xlsx')

os.rename(tempfilename, result_filename)

